#!/usr/bin/env python3
"""Generate readable XLSX from CCaaS dashboard data dictionary content."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "ccaas-contact-center-dashboard-data-dictionary.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
TITLE_FONT = Font(bold=True, size=14)
SUB_FONT = Font(size=10, italic=True, color="44546A")
THIN = Side(style="thin", color="B4B4B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header_row(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER


def write_table(ws, start_row: int, headers: list[str], rows: list[list[str]], col_widths: list[float | None]):
    ncols = len(headers)
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, start_row, ncols)
    for i, row in enumerate(rows, start=start_row + 1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER
    for j, w in enumerate(col_widths, start=1):
        if w:
            ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(ncols)}{start_row + len(rows)}"
    return start_row + len(rows) + 2


def main():
    wb = Workbook()

    # --- Read me ---
    ws0 = wb.active
    ws0.title = "Read me"
    ws0["A1"] = "CCaaS Contact Center Dashboard — Data Dictionary"
    ws0["A1"].font = TITLE_FONT
    ws0["A3"] = "Reference HTML:"
    ws0["B3"] = "ccaas-contact-center-dashboard.html"
    ws0["A4"] = "Purpose:"
    ws0["B4"] = (
        "Align labels, wire APIs, and change definitions consistently. "
        "Each sheet is one section of the dictionary."
    )
    ws0["A6"] = "Notes:"
    ws0["A7"] = (
        "• KPI formulas follow common CCaaS / WFM practice; adapt denominators to your platform.\n"
        "• Demo dashboard: only some KPIs jitter on Refresh; see Shell sheet for caveat."
    )
    ws0["A7"].alignment = WRAP
    for r in (3, 4, 7):
        ws0.row_dimensions[r].height = None
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 85

    # --- Legend ---
    ws = wb.create_sheet("Legend", 1)
    headers = ["Column", "Meaning"]
    data = [
        ["Widget / attribute", "Name shown in UI or code identifier"],
        ["Type", "UI control, KPI card, chart, table column, etc."],
        ["Description", "What it represents"],
        ["Formula / definition", "How to compute or source it (industry-standard where applicable)"],
        ["Why it’s shown", "Operational or analytic purpose"],
    ]
    write_table(ws, 1, headers, data, [28, 70])

    # --- Shell ---
    ws = wb.create_sheet("1 Shell & filters", 2)
    headers = ["Widget / attribute", "Type", "Description", "Formula / definition", "Why it’s shown"]
    data = [
        [
            "Nextiva NCC / AI Dashboard / Contact Center Analytics",
            "Brand block",
            "Product framing; not a metric",
            "Static copy",
            "Orientation and product context",
        ],
        ["Executive", "Nav item", "Executive summary view (active in demo)", "Route / view id (not wired)", "Entry point for leadership KPIs"],
        ["Voice queues", "Nav item", "Queue-centric analytics", "Route / view id", "Drill-down to queue SL, ASA, staffing"],
        ["IVR & self‑serve", "Nav item", "IVR containment, deflection, abandon", "Route / view id", "Isolate automation vs agent demand"],
        ["Agent performance", "Nav item", "Per-agent or team productivity", "Route / view id", "Coaching and capacity planning"],
        ["CSAT & QA", "Nav item", "Quality and satisfaction", "Route / view id", "Outcome quality beyond efficiency"],
        ["Realtime wallboard", "Nav item", "Live operational snapshot", "Route / view id", "Intraday reaction to spikes"],
        ["Contact center operations", "Page title", "Main content heading", "Static", "Describes the working view"],
        [
            "Clock (#clock)",
            "Live text",
            "Current local date/time",
            "new Date() formatted each second",
            "Audit “as of” for screenshots and ops",
        ],
        [
            "Range (#range): Today / Last 7 days / Last 30 days",
            "Filter",
            "Reporting window",
            "Apply as WHERE event_ts in [start, end]",
            "Trends vs intraday; comparisons should use same window",
        ],
        [
            "Queue (#queue): All / Sales / Support / Billing",
            "Filter",
            "Subset of queues/skills",
            "Filter fact tables by queue_id or skill",
            "Focused triage without rebuilding the page",
        ],
        [
            "Refresh (#refresh)",
            "Button",
            "Re-randomizes subset of KPIs and rebuilds charts (demo)",
            "Demo: jitterMetrics() + buildCharts(); prod: refetch API/cache",
            "Fresh aggregates",
        ],
    ]
    next_r = write_table(ws, 1, headers, data, [34, 14, 28, 38, 28])
    ws.cell(row=next_r, column=1, value="Demo caveat:")
    ws.cell(row=next_r, column=1).font = Font(bold=True)
    ws.cell(
        row=next_r,
        column=2,
        value="Changing Range/Queue only runs jitterMetrics() on four count KPIs; charts and most KPI text do not re-query until backend logic exists.",
    )
    ws.merge_cells(start_row=next_r, start_column=2, end_row=next_r, end_column=5)
    ws.cell(row=next_r, column=2).alignment = WRAP
    ws.cell(row=next_r, column=2).font = BODY_FONT

    # --- KPI ---
    ws = wb.create_sheet("2 KPI grid", 3)
    headers = ["Label", "data-metric", "Type", "Description", "Formula / definition", "Why it’s shown"]
    data = [
        [
            "Calls offered",
            "offered",
            "KPI count",
            "Inbound attempts presented to routing (IVR or queue), per platform definition",
            "COUNT(*) where call_direction=inbound AND presented_to_platform=true in period; exclude test trunks if applicable",
            "Top-of-funnel volume for capacity and forecasting",
        ],
        [
            "Calls handled",
            "handled",
            "KPI count",
            "Calls completed with handled disposition (answered + wrapped per policy)",
            "COUNT(*) where final_disposition IN ('answered','handled'); align with AHT denominator",
            "Throughput and staffing adequacy",
        ],
        [
            "Missed calls",
            "missed",
            "KPI count",
            "Offered but not answered within policy (e.g. ring-no-answer, agent timeout)",
            "Platform-specific; often offered_to_agent AND NOT answered; define vs busy/reject",
            "Answerability gaps",
        ],
        [
            "Voicemails",
            "vm",
            "KPI count",
            "Calls leaving voicemail / async capture",
            "COUNT(*) where disposition=voicemail (or VM deposit)",
            "Callback workload and customer preference",
        ],
        [
            "IVR contained",
            "ivr",
            "KPI %",
            "Self-serve in IVR without queueing to agent",
            "100 × (IVR_contained_calls / IVR_presented_calls); define “contained”",
            "Self-service effectiveness; lowers agent load",
        ],
        [
            "IVR abandon",
            "ivrAbandon",
            "KPI %",
            "Disconnect while still in IVR (before agent)",
            "100 × (IVR_abandon_calls / IVR_presented_calls)",
            "IVR friction, prompts, or misrouted intents",
        ],
        [
            "Queue abandoned",
            "abandon",
            "KPI count",
            "Disconnect while waiting for agent (post-IVR)",
            "COUNT(*) where disposition=abandoned_in_queue AND wait_start before disconnect",
            "Staffing and SL pain point",
        ],
        [
            "Abandon rate",
            "abandonRate",
            "KPI %",
            "Share of queue-eligible calls that abandon",
            "100 × queue_abandons / calls_entered_queue OR 100 × abandons / (answered + abandons) — pick one",
            "Normalize abandon vs volume across intervals",
        ],
        [
            "Service level (20s)",
            "sl",
            "KPI %",
            "% answered within threshold",
            "100 × (calls_answered_within_20s / calls_offered_to_queue); threshold must match SLA",
            "Classic SLA for clients and WFM",
        ],
        [
            "ASA",
            "asa",
            "KPI duration",
            "Mean wait in queue before answer",
            "SUM(queue_wait_seconds_answered) / COUNT(answered_calls); often excludes abandons",
            "Speed-of-answer; pairs with SL",
        ],
        [
            "AHT",
            "aht",
            "KPI duration",
            "Average handle time per handled interaction",
            "SUM(talk + hold + ACW) / COUNT(handled_calls); ACW per policy",
            "Efficiency and scheduling (Erlang)",
        ],
        [
            "FCR",
            "fcr",
            "KPI %",
            "Resolved on first contact (no repeat within window)",
            "100 × (no_repeat_within_N_days / eligible_contacts); fix N and channel rules",
            "Quality and cost avoidance",
        ],
        [
            "Occupancy",
            "occ",
            "KPI %",
            "Time agents busy vs available",
            "100 × productive_time / (productive_time + available_time); exclude approved offline",
            "Utilization without burnout",
        ],
        [
            "Adherence",
            "adh",
            "KPI %",
            "Conformance to schedule",
            "100 × minutes_in_schedule_states / scheduled_minutes (per WFM breaks rules)",
            "Schedule discipline for SL",
        ],
        [
            "Callback kept",
            "cb",
            "KPI %",
            "Scheduled callbacks completed vs offered",
            "100 × (callbacks_connected_or_resolved / callbacks_scheduled)",
            "Promise reliability",
        ],
        [
            "CSAT",
            "csat",
            "KPI score",
            "Customer satisfaction (survey)",
            "AVG(survey_score) on agreed scale (here 1–5)",
            "Outcome beyond operational KPIs",
        ],
        [
            "KPI delta line",
            "(subtext)",
            "Subtext",
            "Trend vs prior window",
            "(current − prior) / prior × 100 for counts/rates; or ppt for percentage points",
            "Quick directionality",
        ],
    ]
    r_end = write_table(ws, 1, headers, data, [22, 12, 12, 32, 42, 26])
    ws.cell(row=r_end, column=1, value="Demo caveat:")
    ws.cell(row=r_end, column=1).font = Font(bold=True)
    ws.cell(
        row=r_end,
        column=2,
        value="Only offered, handled, missed, vm jitter on Refresh/Range/Queue. Other KPIs and deltas are static until data-backed.",
    )
    ws.merge_cells(start_row=r_end, start_column=2, end_row=r_end, end_column=6)
    ws.cell(row=r_end, column=2).alignment = WRAP

    def chart_sheet(title: str, chart_id: str, headers, rows, note: str):
        w = wb.create_sheet(title)
        w["A1"] = chart_id
        w["A1"].font = SUB_FONT
        note_row = write_table(w, 3, headers, rows, [22, 14, 28, 42, 26])
        w.cell(row=note_row, column=1, value="Notes:")
        w.cell(row=note_row, column=1).font = Font(bold=True)
        w.cell(row=note_row, column=2, value=note)
        w.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=5)
        w.cell(row=note_row, column=2).alignment = WRAP

    h5 = ["Attribute", "Type", "Description", "Formula / definition", "Why it’s shown"]
    chart_sheet(
        "3a Outcome mix",
        "Call outcome mix (#chartOutcomes) — doughnut",
        h5,
        [
            ["(Widget)", "Doughnut chart", "Distribution of disposition buckets", "Each slice = category count; define mutual exclusivity", "Balance of success vs leakage"],
            ["Handled", "Series segment", 'Same as KPI "Calls handled"', "Count of handled", "Success mix"],
            ["Missed", "Series segment", "Same as Missed calls", "Count", "Answer gap"],
            ["Voicemail", "Series segment", "Same as Voicemails", "Count", "Async workload"],
            ["IVR contained", "Series segment", "Resolved in IVR", "Count (subset of IVR presented)", "Automation value"],
            ["Queue abandon", "Series segment", "Abandoned waiting for agent", "Count", "Queue pain"],
            ["IVR abandon", "Series segment", "Abandoned in IVR", "Count", "IVR UX / routing"],
        ],
        "Demo data: [11204, 412, 189, 3200, 623, 180]. Totals need not equal offered unless you enforce one partition.",
    )
    chart_sheet(
        "3b Hourly volume",
        "Hourly contact volume (#chartVolume) — line",
        h5,
        [
            ["X-axis", "Time bucket", "e.g. 8:00–19:00 hourly", "GROUP BY hour(local_ts)", "Intraday staffing"],
            ["Offered", "Line series", "Calls offered per hour", "COUNT(offered) per hour", "Demand curve"],
            ["Answered", "Line series", "Calls answered per hour", "COUNT(answered) per hour", "Served vs offered"],
        ],
        "Demo: 12 hard-coded hourly points; replace with query for selected Range.",
    )
    chart_sheet(
        "3c IVR funnel",
        "IVR funnel (#chartIvr) — bar",
        h5,
        [
            ["Presented", "Bar", "Calls that hit IVR", "COUNT(IVR_session_start)", "Funnel top"],
            ["Contained", "Bar", "Self-serve completion", "COUNT(contained_disposition)", "Deflection"],
            ["To agent", "Bar", "Transferred/queued to agent", "COUNT(IVR_exit_to_queue)", "Agent demand from IVR"],
            ["IVR abandon", "Bar", "Disconnect in IVR", "COUNT(abandon_in_IVR)", "Drop-off"],
            ["Error / zero-out", "Bar", "Failures or operator request", "COUNT(error OR zero_out)", "Tree health"],
        ],
        "Demo: [12847, 8787, 2850, 655, 180]. Define session-level vs event-level counting.",
    )
    chart_sheet(
        "3d Queue health",
        "Queue health (#chartQueues) — grouped bar",
        h5,
        [
            ["Categories", "X-axis", "Queue names (Sales, Support, Billing, VIP)", "One group per queue_id", "Compare queues"],
            ["ASA (sec)", "Bar (left Y)", "ASA for that queue", "ASA formula filtered by queue", "Wait by line of business"],
            ["Abandon %", "Bar (right Y)", "Abandon rate for that queue", "100 × abandons / (answered + abandons) or your standard", "Pain vs ASA"],
        ],
        "Demo: ASA [38,65,22,18]s; Abandon % [4.2,7.1,3.0,1.2]. SL by queue is not a series in this chart.",
    )

    ws = wb.create_sheet("4 Queue snapshot")
    headers = ["Column", "Type", "Description", "Formula / definition", "Why it’s shown"]
    data = [
        ["Queue", "Text", "Queue or virtual queue name", "Dimension queue.name", "Identity"],
        ["Waiting", "Number or —", "Current calls in queue", "COUNT(calls WHERE state=waiting) realtime", "Immediate backlog"],
        ["Longest wait", "Duration or text", "Oldest waiting call age", "MAX(now - entered_queue_ts); or SLA text for callback", "Worst-case wait"],
        ["Agents avail", "Number or —", "Agents ready for next call", "COUNT(agents WHERE state=available)", "Immediate capacity"],
        ["ASA (rolling)", "Duration or —", "Rolling avg wait for recent answered", "Mean over last N minutes or M calls", "Short-term wallboard trend"],
        ["Abandon %", "Percent", "Rolling or interval abandon rate", "Same abandon % on rolling window", "Risk if climbing"],
        ["Type", "Tag (VOICE/IVR/QUEUE)", "Channel or object kind", "Categorical", "Drill-down mental model"],
    ]
    next_r = write_table(ws, 1, headers, data, [18, 16, 30, 40, 28])
    ws.cell(row=next_r, column=1, value="Demo:")
    ws.cell(row=next_r, column=2, value="Hard-coded queueRows in HTML script; not live.")
    ws.merge_cells(start_row=next_r, start_column=2, end_row=next_r, end_column=5)

    ws = wb.create_sheet("5 Consistency")
    headers = ["Topic", "Recommendation"]
    data = [
        ["Denominator for SL / abandon", "Use one canonical “calls offered to queue” definition everywhere."],
        ["IVR vs queue", "Split by timestamps: before queue_enter vs after, for IVR vs queue abandon."],
        ["Handled vs answered", "State if AHT and “handled” require ACW complete."],
        ["Duplicate sessions", "Define if abandon then recall counts as one or two offered."],
    ]
    write_table(ws, 1, headers, data, [32, 70])

    ws = wb.create_sheet("6 Metric index")
    headers = ["data-metric", "KPI label"]
    data = [
        ["offered", "Calls offered"],
        ["handled", "Calls handled"],
        ["missed", "Missed calls"],
        ["vm", "Voicemails"],
        ["ivr", "IVR contained"],
        ["ivrAbandon", "IVR abandon"],
        ["abandon", "Queue abandoned"],
        ["abandonRate", "Abandon rate"],
        ["sl", "Service level (20s)"],
        ["asa", "ASA"],
        ["aht", "AHT"],
        ["fcr", "FCR"],
        ["occ", "Occupancy"],
        ["adh", "Adherence"],
        ["cb", "Callback kept"],
        ["csat", "CSAT"],
    ]
    write_table(ws, 1, headers, data, [18, 32])

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
